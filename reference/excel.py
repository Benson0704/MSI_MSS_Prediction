import openpyxl
total = {'20_crop': [], '72_crop': [], '81_crop': []}
li = ['20_crop', '72_crop', '81_crop']
f = open('new.txt', 'r')
for i in f.readlines():
    a = i.replace('\n', '')
    total[a.split('_&_')[0]].append((a.split(' ')
                                     [0].split('_&_')[1], a.split(' ')[-1]))
f.close()
wb = openpyxl.Workbook()
for i in range(1, 4):
    wb.create_sheet(index=i-1, title=li[i-1])
    sheet = wb.worksheets[i-1]
    for j, data in enumerate(total[li[i-1]]):
        a, b = data
        sheet.cell(j+1, 1, a)
        sheet.cell(j+1, 2, b)
wb.save('1.xlsx')
total = {'20_crop': [], '72_crop': [], '81_crop': []}
li = ['20_crop', '72_crop', '81_crop']
f = open('new2.txt', 'r')
for i in f.readlines():
    a = i.replace('\n', '')
    total[a.split('_&_')[0]].append((a.split(' ')
                                     [0].split('_&_')[1], a.split(' ')[1], a.split(' ')[2], a.split(' ')[3]))
f.close()
wb = openpyxl.Workbook()
for i in range(1, 4):
    wb.create_sheet(index=i-1, title=li[i-1])
    sheet = wb.worksheets[i-1]
    for j, data in enumerate(total[li[i-1]]):
        a, b, c, d = data
        sheet.cell(j+1, 1, a)
        sheet.cell(j+1, 2, b)
        sheet.cell(j+1, 3, c)
        sheet.cell(j+1, 4, d)
wb.save('2.xlsx')
