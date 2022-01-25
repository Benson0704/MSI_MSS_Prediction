import random
from sklearn import linear_model
from sklearn import neighbors
from sklearn import metrics
from sklearn.model_selection import cross_val_score
from sklearn import svm
import time
import openpyxl
import matplotlib.pyplot
import numpy
import scipy
DO_RAW = 1
DO_LR = 1
DO_SVC = 0
DO_KNN = 0
LR_C = 0.0699


def paint_roc(label_list, predicted_list, output_str='test.png'):
    fpr, tpr, thersholds = metrics.roc_curve(
        label_list, predicted_list, pos_label=1)
    roc_auc = metrics.auc(fpr, tpr)
    result.writelines('{}: auc in roc painting is: {}'.format(
        output_str.split('.')[0], roc_auc))
    matplotlib.pyplot.plot(
        fpr, tpr, '-', label='ROC (area = {0:.4f})'.format(roc_auc), lw=1)

    matplotlib.pyplot.xlim([-0.05, 1.05])  # 设置x、y轴的上下限，以免和边缘重合，更好的观察图像的整体
    matplotlib.pyplot.ylim([-0.05, 1.05])
    matplotlib.pyplot.xlabel('False Positive Rate')
    matplotlib.pyplot.ylabel('True Positive Rate')  # 可以使用中文，但需要导入一些库即字体
    matplotlib.pyplot.title('ROC Curve')
    matplotlib.pyplot.legend(loc="lower right")
    matplotlib.pyplot.savefig(output_str.replace('.png', '.eps'), format='eps')
    matplotlib.pyplot.close()


train = open('train.txt', 'r')
test = open('test.txt', 'r')
result = open('result.txt', 'w')
result.writelines(time.asctime(time.localtime())+' Starts!\n')
train_name = []
train_score = []
train_label = []
for line in train.readlines():
    train_name.append(line[:-1].split(' ')[0].split('/')[-1])
    train_score.append(float(line[:-1].split(' ')[1][:8]))
    train_label.append(int(line[:-1].split(' ')[2]))
train.close()
test_name = []
test_score = []
test_label = []
for line in test.readlines():
    test_name.append(line[:-1].split(' ')[0].split('/')[-1])
    test_score.append(float(line[:-1].split(' ')[1][:8]))
    test_label.append(int(line[:-1].split(' ')[2]))
test.close()


# raw 0.5 acc
if DO_RAW:
    right = 0
    for i, _ in enumerate(test_label):
        right += ((test_score[i] >= 0.5) == test_label[i])
    raw_result = [int(x >= 0.5) for x in test_score]
    result.writelines('raw 0.5 acc: {:.6f}\n'.format(right/len(test_label)))
    result.writelines('predict results:\n'+str(raw_result) +
                      '\nwrong cases:(rank, score, label)\n')
    for i, _ in enumerate(test_label):
        if test_label[i] != raw_result[i]:
            result.writelines(str((i, test_score[i], test_label[i]))+', ')
    paint_roc(test_label, test_score, 'raw.png')
    result.writelines('\n')


def cross_validation(train_data, label):
    train_data0 = train_data[0:len(train_data)//5]
    train_data1 = train_data[len(train_data)//5:len(train_data)//5*2]
    train_data2 = train_data[len(train_data)//5*2:len(train_data)//5*3]
    train_data3 = train_data[len(train_data)//5*3:len(train_data)//5*4]
    train_data4 = train_data[len(train_data)//5*4:]
    train_datas = [train_data0, train_data1,
                   train_data2, train_data3, train_data4]
    label0 = label[0:len(label)//5]
    label1 = label[len(label)//5:len(label)//5*2]
    label2 = label[len(label)//5*2:len(label)//5*3]
    label3 = label[len(label)//5*3:len(label)//5*4]
    label4 = label[len(label)//5*4:]
    labels = [label0, label1,
              label2, label3, label4]
    result = []
    for i in range(5):
        train_CV = []
        test_CV = []
        for x in range(5):
            if x != i:
                train_CV += train_datas[x]
            else:
                test_CV += train_datas[x]
        train_label = []
        test_label = []
        for x in range(5):
            if x != i:
                train_label += labels[x]
            else:
                test_label += labels[x]
        max_acc = 0
        max_acc_prob = 0
        for j in range(0, 1000):
            train_correct = 0
            train_total = 0
            train_result = []
            for x in train_CV:
                train_result.append(x > j*0.001)
            for x, _ in enumerate(train_result):
                train_total += 1
                train_correct += (train_result[x] == train_label[x])
            if train_correct/train_total > max_acc:
                max_acc = train_correct/train_total
                max_acc_prob = j
        test_correct = 0
        test_total = 0
        test_result = []
        for x in test_CV:
            test_result.append(x > max_acc_prob*0.001)
        for x, _ in enumerate(test_result):
            test_total += 1
            test_correct += (test_result[x] == test_label[x])
        result.append(max_acc_prob*0.001)
        print('{}: prob: {} acc: {}'.format(
            i, max_acc_prob*0.001, test_correct/test_total))
    return result


max_acc = 0
max_acc_prob = 0
r = cross_validation(train_score, train_label)
for j in range(0, 1000):
    train_correct = 0
    train_total = 0
    train_result = []
    for x in train_score:
        train_result.append(x > j*0.001)
    for x, _ in enumerate(train_result):
        train_total += 1
        train_correct += (train_result[x] == train_label[x])
    if train_correct/train_total > max_acc:
        max_acc = train_correct/train_total
        max_acc_prob = j
print('max acc on train: ', max_acc, 'divided at: ', max_acc_prob*0.001)

print('0.500')
test_correct = 0
test_total = 0
test_result = []
for x in test_score:
    test_result.append(x > 500*0.001)
for x, _ in enumerate(test_result):
    test_total += 1
    test_correct += (test_result[x] == test_label[x])
print('test acc: ', test_correct/test_total)
print('result: ', test_result)
print('label: ', test_label)
print('score: ', test_score)
real0to0 = 0
real0to1 = 0
real1to0 = 0
real1to1 = 0
for x, _ in enumerate(test_result):
    if test_label[x] == 0:
        if test_result[x] == 0:
            real0to0 += 1
        else:
            real0to1 += 1
    else:
        if test_result[x] == 0:
            real1to0 += 1
        else:
            real1to1 += 1
print(
    '   predict0 predict1\nreal0 {} {}\nreal1 {} {}'.format(
        real0to0, real0to1, real1to0, real1to1)
)
acc = []
for i in range(500):
    pt = []
    plt = []
    test_correct = 0
    test_total = 0
    for j in range(len(test_label)):
        x = random.randint(0, len(test_label)-1)
        pt.append(test_result[x])
        plt.append(test_label[x])
    for x, _ in enumerate(pt):
        test_total += 1
        test_correct += (pt[x] == plt[x])
    acc.append(test_correct/test_total)
acc = numpy.array(acc)
mean, std = acc.mean(), acc.std(ddof=1)
print('(bootstrap 500) mean is ' + str(mean))
conf_intveral = scipy.stats.norm.interval(0.95, loc=mean, scale=std)
print('conf_intveral ' + str(conf_intveral))

# 最重要的函数：通过precision_recall_curve()函数，求出recall，precision，以及阈值
precision, recall, thresholds = metrics.precision_recall_curve(
    test_label, test_score)
print(precision)
print(recall)
matplotlib.pyplot.plot(recall, precision, lw=1)
matplotlib.pyplot.plot([0, 1], [0, 1], '--',
                       color=(0.6, 0.6, 0.6), label="Luck")				 # 画对角线
matplotlib.pyplot.xlim([-0.05, 1.05])
matplotlib.pyplot.ylim([-0.05, 1.05])
matplotlib.pyplot.xlabel("Recall Rate")
matplotlib.pyplot.ylabel("Precision Rate")
matplotlib.pyplot.savefig('pr_0.5.eps')
print('0.307')
test_correct = 0
test_total = 0
test_result = []
for x in test_score:
    test_result.append(x > max_acc_prob*0.001)
for x, _ in enumerate(test_result):
    test_total += 1
    test_correct += (test_result[x] == test_label[x])
print('test acc: ', test_correct/test_total)
print('result: ', test_result)
print('label: ', test_label)
print('score: ', test_score)
real0to0 = 0
real0to1 = 0
real1to0 = 0
real1to1 = 0
for x, _ in enumerate(test_result):
    if test_label[x] == 0:
        if test_result[x] == 0:
            real0to0 += 1
        else:
            real0to1 += 1
    else:
        if test_result[x] == 0:
            real1to0 += 1
        else:
            real1to1 += 1
print(
    '   predict0 predict1\nreal0 {} {}\nreal1 {} {}'.format(
        real0to0, real0to1, real1to0, real1to1)
)
acc = []
for i in range(500):
    pt = []
    plt = []
    test_correct = 0
    test_total = 0
    for j in range(len(test_label)):
        x = random.randint(0, len(test_label)-1)
        pt.append(test_result[x])
        plt.append(test_label[x])
    for x, _ in enumerate(pt):
        test_total += 1
        test_correct += (pt[x] == plt[x])
    acc.append(test_correct/test_total)
acc = numpy.array(acc)
mean, std = acc.mean(), acc.std(ddof=1)
print('(bootstrap 500) mean is ' + str(mean))
conf_intveral = scipy.stats.norm.interval(0.95, loc=mean, scale=std)
print('conf_intveral ' + str(conf_intveral))
precision, recall, thresholds = metrics.precision_recall_curve(
    test_label, test_score)
print(precision)
print(recall)
matplotlib.pyplot.plot(recall, precision, lw=1)
matplotlib.pyplot.plot([0, 1], [0, 1], '--',
                       color=(0.6, 0.6, 0.6), label="Luck")				 # 画对角线
matplotlib.pyplot.xlim([-0.05, 1.05])
matplotlib.pyplot.ylim([-0.05, 1.05])
matplotlib.pyplot.xlabel("Recall Rate")
matplotlib.pyplot.ylabel("Precision Rate")
matplotlib.pyplot.savefig('pr_0.307.eps')
# LR
if DO_LR:
    LR_acc = 0
    LR_train_data = [[x] for x in train_score]
    LR_train_label = [x for x in train_label]
    if 'LR_C' not in locals().keys():
        LR_C = 0
        for iter in range(1, 2000):
            LR = linear_model.LogisticRegression(C=0.0001*iter)

            LR_score = cross_val_score(LR, LR_train_data, LR_train_label)
            LR_score = list(LR_score)
            LR_tmp = 0
            for x in LR_score:
                LR_tmp += x
            LR_tmp /= 5
            '''
        LR.fit(LR_train_data, LR_train_label)
        LR_result = list(LR.predict([[x] for x in train_score]))
        right = 0
        for i, _ in enumerate(LR_train_label):
                right += (LR_result[i] == LR_train_label[i])
            LR_tmp = right/len(LR_train_label)
            '''
            if LR_tmp > LR_acc:
                LR_acc = LR_tmp
                LR_C = iter*0.0001
            if iter % 20 == 0:
                print('LR', iter)
        result.writelines(
            '\nLR: validation acc is {}, LR_C: {}\n'.format(LR_acc, LR_C))
    LR = linear_model.LogisticRegression(C=LR_C)
    LR.fit(LR_train_data, LR_train_label)
    LR_result = list(LR.predict([[x] for x in test_score]))
    LR_probability = LR.predict_proba([[x] for x in test_score])
    probability = []
    for i in LR_probability:
        probability.append(i[1])
    right = 0
    for i, _ in enumerate(test_label):
        right += (LR_result[i] == test_label[i])
    result.writelines('LR acc: {:.6f}\n'.format(right/len(test_label)))
    result.writelines('predict results:\n'+str(LR_result) +
                      '\nwrong cases:(rank, score, label)\n')
    for i, _ in enumerate(test_label):
        if test_label[i] != LR_result[i]:
            result.writelines(str((i, test_score[i], test_label[i]))+', ')
    paint_roc(test_label, probability, 'LR1.png')
    paint_roc(test_label, test_score, 'LR2.png')
    paint_roc(test_label, [random.random()
                           for _ in probability], 'LR3.png')
    paint_roc(test_label, LR_result, 'LR4.png')
    print('intercept:', LR.intercept_, 'coef:', LR.coef_)
    print('score:\n', test_score)
    print('LR result:\n', probability)

    result.writelines('\n')

# SVM-SVC
if DO_SVC:
    SVC_C = 0
    SVC_acc = 0
    SVC_train_data = [[x] for x in train_score]
    SVC_train_label = [x for x in train_label]

    for iter in range(1, 2000):
        SVC = svm.SVC(C=0.001*iter)

        SVC_score = cross_val_score(SVC, SVC_train_data, SVC_train_label)
        SVC_score = list(SVC_score)
        SVC_tmp = 0
        for x in SVC_score:
            SVC_tmp += x
        SVC_tmp /= 5
        '''
        SVC.fit(SVC_train_data, SVC_train_label)
        SVC_result = list(SVC.predict([[x] for x in train_score]))
        right = 0
        for i, _ in enumerate(SVC_train_label):
            right += (SVC_result[i] == SVC_train_label[i])
        SVC_tmp = right/len(SVC_train_label)
        '''
        if SVC_tmp >= SVC_acc:
            SVC_acc = SVC_tmp
            SVC_C = iter*0.001
        if iter % 20 == 0:
            print('SVC', iter)
    result.writelines(
        '\nSVC: validation acc is {}, SVC_C: {}\n'.format(SVC_acc, SVC_C))
    SVC = svm.SVC(C=1)
    SVC.fit(SVC_train_data, SVC_train_label)
    SVC_result = list(SVC.predict([[x] for x in test_score]))
    right = 0
    for i, _ in enumerate(test_label):
        right += (SVC_result[i] == test_label[i])
    result.writelines('SVC acc: {:.6f}\n'.format(right/len(test_label)))
    result.writelines('predict results:\n'+str(SVC_result) +
                      '\nwrong cases:(rank, score, label)\n')
    for i, _ in enumerate(test_label):
        if test_label[i] != SVC_result[i]:
            result.writelines(str((i, test_score[i], test_label[i]))+', ')
    paint_roc(test_label, SVC_result, 'SVC.png')
    result.writelines('\n')

# KNN
if DO_KNN:
    KNN_k = 0
    KNN_acc = 0
    KNN_train_data = [[x] for x in train_score]
    KNN_train_label = [x for x in train_label]

    for iter in range(1, 10):
        KNN = neighbors.KNeighborsClassifier(n_neighbors=iter)

        KNN_score = cross_val_score(KNN, KNN_train_data, KNN_train_label)
        KNN_score = list(KNN_score)
        KNN_tmp = 0
        for x in KNN_score:
            KNN_tmp += x
        KNN_tmp /= 5
        '''
        KNN.fit(KNN_train_data, KNN_train_label)
        KNN_result = list(KNN.predict([[x] for x in train_score]))
        right = 0
        for i, _ in enumerate(KNN_train_label):
            right += (KNN_result[i] == KNN_train_label[i])
        KNN_tmp = right/len(KNN_train_label)
        '''
        if KNN_tmp >= KNN_acc:
            KNN_acc = KNN_tmp
            KNN_k = iter
    result.writelines(
        '\nKNN: validation acc is {}, KNN_k: {}\n'.format(KNN_acc, KNN_k))
    KNN = neighbors.KNeighborsClassifier(n_neighbors=5)
    KNN.fit(KNN_train_data, KNN_train_label)
    KNN_result = list(KNN.predict([[x] for x in test_score]))
    right = 0
    for i, _ in enumerate(test_label):
        right += (KNN_result[i] == test_label[i])
    result.writelines('KNN acc: {:.6f}\n'.format(right/len(test_label)))
    result.writelines('predict results:\n'+str(KNN_result) +
                      '\nwrong cases:(rank, score, label)\n')
    for i, _ in enumerate(test_label):
        if test_label[i] != KNN_result[i]:
            result.writelines(str((i, test_score[i], test_label[i]))+', ')
    paint_roc(test_label, KNN_result, 'KNN.png')
    result.writelines('\n')


result.writelines('\nothers\n')
result.writelines('test_patients labels: \n' + str(test_label)+'\n')
result.writelines('test_patients names:\n '+str(test_name)+'\n')
result.writelines('test_patients scores:\n ' + str(test_score)+'\n')
result.close()

wb = openpyxl.Workbook()
wb.create_sheet('之前的测试数据(三月份)')
sheet = wb.worksheets[-1]
sheet.cell(1, 1, '88 patients, MSS=1, MSI=0')
sheet.cell(1, 2, 'label')
sheet.cell(1, 3, 'score')
if DO_LR:
    sheet.cell(1, 4, 'LR predict')
if DO_SVC:
    sheet.cell(1, 5, 'SVC predict')
if DO_KNN:
    sheet.cell(1, 6, 'KNN predict')
if DO_RAW:
    sheet.cell(1, 7, 'raw predict')
for i in range(len(test_score)):
    sheet.cell(i+2, 1, test_name[i])
    sheet.cell(i+2, 2, test_label[i])
    sheet.cell(i+2, 3, test_score[i])
    if DO_LR:
        sheet.cell(i+2, 4, LR_result[i])
    if DO_SVC:
        sheet.cell(i+2, 5, SVC_result[i])
    if DO_KNN:
        sheet.cell(i+2, 6, KNN_result[i])
    if DO_RAW:
        sheet.cell(i+2, 7, raw_result[i])
wb.save('2021.3data.xlsx')
"""

newfile = open('MSS.txt', 'r')
patients = {}
MSI_group = os.listdir('newdata/MSIMSS-1')
MSS_group = os.listdir('newdata/MSIMSS-2')
MSS_file = open('mss.txt', 'w')
MSI_file = open('msi.txt', 'w')
MSS_name = []
MSI_name = []
MSS_score = []
MSI_score = []
MSS_label = []
MSI_label = []
MSS_num = []
MSI_num = []
for line in newfile.readlines():
    if line.split('/')[1] not in patients.keys():
        patients[line.split('/')[1]] = [float(line.split(' ')[1]) > 0.5, 1]
    else:
        patients[line.split('/')[1]][1] += 1
        patients[line.split('/')[1]][0] += float(line.split(' ')[1]) > 0.5
for k, v in patients.items():
    tmp = patients[k]
    patients[k] = [tmp[0]/tmp[1]]
    if k in MSI_group:
        MSI_num.append(tmp[1])
        patients[k].append(0)
        MSI_file.writelines(
            k+' '+str(patients[k][0])+' '+str(patients[k][1])+'\n')
        MSI_name.append(k)
        MSI_score.append(patients[k][0])
        MSI_label.append(patients[k][1])
    elif k in MSS_group:
        MSS_num.append(tmp[1])
        patients[k].append(1)
        MSS_file.writelines(
            k+' '+str(patients[k][0])+' '+str(patients[k][1])+'\n')
        MSS_name.append(k)
        MSS_score.append(patients[k][0])
        MSS_label.append(patients[k][1])
    else:
        print('exit')

print(patients)
wb = openpyxl.Workbook()

LR_C = 0.0699
LR_train_data = [[x] for x in train_score]
LR_train_label = [x for x in train_label]
LR = linear_model.LogisticRegression(C=LR_C)
LR.fit(LR_train_data, LR_train_label)
LR_result = list(LR.predict([[x] for x in MSI_score]))
right = 0
for i, _ in enumerate(MSI_label):
    right += (LR_result[i] == MSI_label[i])
result.writelines('MSI LR acc: {:.6f}\n'.format(right/len(MSI_label)))
result.writelines('MSI predict results:\n'+str(LR_result) +
                  '\nwrong cases:(rank, score, label)\n')
for i, _ in enumerate(MSI_label):
    if MSI_label[i] != LR_result[i]:
        result.writelines(str((i, MSI_score[i], MSI_label[i]))+', ')
paint_roc(MSI_label, LR_result, 'MSI LR.png')
result.writelines('\n')

wb.create_sheet('2021.6.MSIMSS-1')
sheet = wb.worksheets[-1]
sheet.cell(1, 1, 'patients: {}, MSS=1, MSI=0'.format(len(MSI_name)))
sheet.cell(1, 2, 'label')
sheet.cell(1, 3, 'score')
sheet.cell(1, 4, 'LR predict')
sheet.cell(1, 5, 'tile num')
for i in range(len(MSI_score)):
    sheet.cell(i+2, 1, MSI_name[i])
    sheet.cell(i+2, 2, MSI_label[i])
    sheet.cell(i+2, 3, MSI_score[i])
    sheet.cell(i+2, 4, LR_result[i])
    sheet.cell(i+2, 5, MSI_num[i])

LR_train_data = [[x] for x in train_score]
LR_train_label = [x for x in train_label]
LR = linear_model.LogisticRegression(C=LR_C)
LR.fit(LR_train_data, LR_train_label)
LR_result = list(LR.predict([[x] for x in MSS_score]))
right = 0
for i, _ in enumerate(MSS_label):
    right += (LR_result[i] == MSS_label[i])
result.writelines('MSS LR acc: {:.6f}\n'.format(right/len(MSS_label)))
result.writelines('MSS predict results:\n'+str(LR_result) +
                  '\nwrong cases:(rank, score, label)\n')
for i, _ in enumerate(MSS_label):
    if MSS_label[i] != LR_result[i]:
        result.writelines(str((i, MSS_score[i], MSS_label[i]))+', ')
paint_roc(MSS_label, LR_result, 'MSS LR.png')
result.writelines('\n')

wb.create_sheet('2021.6.MSIMSS-2')
sheet = wb.worksheets[-1]
sheet.cell(1, 1, 'patients: {}, MSS=1, MSI=0'.format(len(MSS_name)))
sheet.cell(1, 2, 'label')
sheet.cell(1, 3, 'score')
sheet.cell(1, 4, 'LR predict')
sheet.cell(1, 5, 'tile num')
for i in range(len(MSS_score)):
    sheet.cell(i+2, 1, MSS_name[i])
    sheet.cell(i+2, 2, MSS_label[i])
    sheet.cell(i+2, 3, MSS_score[i])
    sheet.cell(i+2, 4, LR_result[i])
    sheet.cell(i+2, 5, MSS_num[i])
wb.save('2021.6data.xlsx')
"""
